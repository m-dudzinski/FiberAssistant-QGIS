import os
import json
from collections import defaultdict

import os
import json
import re
import csv
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    openpyxl = None

from qgis.PyQt import uic
from qgis.PyQt.QtWidgets import QWidget, QVBoxLayout, QFileDialog
from qgis.PyQt.QtCore import Qt
from qgis.core import (
    QgsProject,
    QgsVectorLayer,
    QgsFeature,
    QgsGeometry,
    QgsWkbTypes,
    QgsPointXY,
    QgsCoordinateReferenceSystem,
    QgsCoordinateTransform,
    QgsFeatureRequest,
    QgsSpatialIndex
)

from ..core.logger import logger
from .base_widget import FormattedOutputWidget

FORM_CLASS, _ = uic.loadUiType(os.path.join(
    os.path.dirname(__file__), '../ui/zarzadzanie_PA_widget.ui'))

class ZarzadzaniePAWidget(QWidget, FORM_CLASS):
    FUNCTIONALITY_NAME = "Zarządzanie PA"

    def __init__(self, iface, parent=None):
        super(ZarzadzaniePAWidget, self).__init__(parent)
        self.iface = iface
        self.logger = logger
        self.setupUi(self)

        self.splitter.setSizes([400, 150])
        self.splitter.setCollapsible(0, False)
        self.splitter.setCollapsible(1, False)

        self._setup_output_widget()
        self._connect_signals()
        self._populate_zakres_combobox()
        self._populate_dzialki_layers_combobox()
        self._setup_initial_state()

    def _get_feature_identifier(self, feature):
        id_budynku = feature.attribute("Id_budynku") or "brak"
        miejscowosc = feature.attribute("Miejscowos") or "brak"
        ulica = feature.attribute("Ulica") or ""
        numer_porz = feature.attribute("Numer porz") or "brak"

        address_parts = [miejscowosc]
        if ulica:
            address_parts.append(ulica)
        address_parts.append(numer_porz)
        
        return f"[Id_budynku: {id_budynku}, Adres: {' '.join(address_parts)} ]"

    def _setup_output_widget(self):
        self.output_widget = FormattedOutputWidget()
        layout = self.output_widget_placeholder.layout()
        if layout is None:
            layout = QVBoxLayout(self.output_widget_placeholder)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.output_widget)
        self.logger.set_user_message_widget(self.output_widget.output_console)

    def _connect_signals(self):
        self.refresh_button.clicked.connect(self.refresh_data)
        # Tab 1
        self.dzialki_layer_combobox.currentIndexChanged.connect(self._on_dzialki_layer_changed)
        self.groupBox_dzialka.toggled.connect(self.formLayout_dzialka.setEnabled)
        self.groupBox_zasilanie.toggled.connect(self.horizontalLayout_zasilanie.setEnabled)
        # Tab 2
        self.browse_button.clicked.connect(self._browse_file)
        # Tab 3
        self.browse_button_mr.clicked.connect(self._browse_file_mr)

    def run_main_action(self):
        self.output_widget.clear_log()
        current_tab_index = self.tabWidget.currentIndex()
        if current_tab_index == 0:
            self.run_przypisania_atrybutow_action()
        elif current_tab_index == 1:
            self.run_wykluczanie_pa_action()
        elif current_tab_index == 2:
            self.run_aktualizuj_mr_action()


    def _populate_zakres_combobox(self):
        self.zakres_combo_box.clear()
        zakres_layer_list = QgsProject.instance().mapLayersByName("zakres_zadania")
        if not zakres_layer_list:
            self.output_widget.log_error("Nie znaleziono warstwy 'zakres_zadania'.")
            return
        if len(zakres_layer_list) > 1:
            self.output_widget.log_warning("Znaleziono więcej niż jedną warstwę o nazwie 'zakres_zadania'. Używana będzie pierwsza z listy.")
        
        zakres_layer = zakres_layer_list[0]

        if "nazwa" not in zakres_layer.fields().names():
            self.output_widget.log_error("Warstwa 'zakres_zadania' nie posiada atrybutu 'nazwa'.")
            return

        features_to_add = []
        for feature in zakres_layer.getFeatures():
            if feature.attribute("nazwa"):
                features_to_add.append((feature["nazwa"], feature.geometry()))

        features_to_add.sort(key=lambda f: f[0])

        for nazwa, geom in features_to_add:
            self.zakres_combo_box.addItem(nazwa, geom)

        self.output_widget.log_info(f"Znaleziono {self.zakres_combo_box.count()} zakresów.")

    def _populate_dzialki_layers_combobox(self):
        self.dzialki_layer_combobox.clear()
        layers = QgsProject.instance().mapLayers().values()
        polygon_layers = [layer for layer in layers if isinstance(layer, QgsVectorLayer) and layer.geometryType() == QgsWkbTypes.PolygonGeometry]
        for layer in polygon_layers:
            self.dzialki_layer_combobox.addItem(layer.name(), layer)
        self._on_dzialki_layer_changed()

    def _on_dzialki_layer_changed(self):
        self.dzialki_attribute_combobox.clear()
        selected_layer = self.dzialki_layer_combobox.currentData()
        if selected_layer:
            self.dzialki_attribute_combobox.addItems([field.name() for field in selected_layer.fields()])

    def _setup_initial_state(self):
        self.formLayout_dzialka.setEnabled(self.groupBox_dzialka.isChecked())
        self.horizontalLayout_zasilanie.setEnabled(self.groupBox_zasilanie.isChecked())

    def refresh_data(self):
        self.output_widget.log_info("Odświeżanie list...")
        self._populate_zakres_combobox()
        self._populate_dzialki_layers_combobox()
        self.output_widget.log_info("Listy zostały zaktualizowane.")

    def _browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Wybierz plik", "", "Pliki Excel (*.xlsx);;Pliki CSV (*.csv)")
        if file_path:
            self.file_path_line_edit.setText(file_path)

    def run_wykluczanie_pa_action(self):
        self.output_widget.log_info("Uruchomiono 'Wykluczanie PA'...")

        pa_layer_list = QgsProject.instance().mapLayersByName("lista_pa")
        if not pa_layer_list:
            self.output_widget.log_error("Nie znaleziono warstwy 'lista_pa'.")
            return
        pa_layer = pa_layer_list[0]

        if not self._validate_wykluczanie_pa(pa_layer):
            self.output_widget.log_error("Walidacja nie powiodła się. Przerwana operacja.")
            return

        self.output_widget.log_success("Walidacja pomyślna. Rozpoczynanie operacji...")

        raw_identifiers, duplicates = self._get_identifiers()
        if not raw_identifiers:
            self.output_widget.log_error("Nie podano żadnych identyfikatorów do przetworzenia.")
            return

        identifiers = {str(id_val).strip() for id_val in raw_identifiers}

        scope_geom = self.zakres_combo_box.currentData()
        overwrite = self.nadpisz_radio_wykluczanie.isChecked()
        target_attr = "X_dubel" if self.radio_do_wykluczenia.isChecked() else "X_weryfik"
        id_attr = "Id_budynku" if self.radio_id_budynku.isChecked() else "X_ID_TOK"
        
        stats = {
            "processed": 0,
            "matched_id_budynku": 0,
            "matched_x_id_tok": 0,
            "x_dubel_set": 0,
            "x_weryfik_set": 0,
            "skipped_existing": 0,
            "skipped_no_geom": 0,
            "skipped_outside_scope": 0,
            "unmatched_ids": set(identifiers),
            "user_duplicates": duplicates
        }

        pa_layer.startEditing()
        request = QgsFeatureRequest().setFilterRect(scope_geom.boundingBox())
        for feature in pa_layer.getFeatures(request):
            stats["processed"] += 1
            pa_geom = feature.geometry()

            if not pa_geom:
                stats["skipped_no_geom"] += 1
                continue
            
            if not pa_geom.intersects(scope_geom):
                stats["skipped_outside_scope"] += 1
                continue

            feature_id_val = str(feature[id_attr] or '').strip()
            if feature_id_val in identifiers:
                if id_attr == "Id_budynku":
                    stats["matched_id_budynku"] += 1
                else:
                    stats["matched_x_id_tok"] += 1
                
                stats["unmatched_ids"].discard(feature_id_val)

                current_value = feature[target_attr]
                if overwrite or not current_value:
                    pa_layer.changeAttributeValue(feature.id(), pa_layer.fields().indexOf(target_attr), "TAK")
                    if target_attr == "X_dubel":
                        stats["x_dubel_set"] += 1
                    else:
                        stats["x_weryfik_set"] += 1
                else:
                    stats["skipped_existing"] += 1

        pa_layer.commitChanges()
        self._log_wykluczanie_summary(stats)

    def run_przypisania_atrybutow_action(self):
        self.output_widget.log_info("Uruchomiono 'Przypisania atrybutów'...")

        pa_layer_list = QgsProject.instance().mapLayersByName("lista_pa")
        if not pa_layer_list:
            self.output_widget.log_error("Nie znaleziono warstwy 'lista_pa'.")
            return
        pa_layer = pa_layer_list[0]

        if not self._is_valid_for_run(pa_layer):
            self.output_widget.log_error("Walidacja nie powiodła się. Przerwana operacja.")
            return

        self.output_widget.log_success("Walidacja pomyślna. Rozpoczynanie operacji...")

        scope_geom = self.zakres_combo_box.currentData()
        overwrite = self.nadpisz_radio.isChecked()
        
        stats = {
            "processed": 0,
            "dzialka_assigned": 0,
            "pe_assigned": 0,
            "md_split_assigned": 0,
            "spl-i-rz_assigned": 0,
            "pe_values": defaultdict(int),
            "md_split_values": defaultdict(int),
            "spl-i-rz_values": defaultdict(int),
            "skipped_no_geom": 0,
            "skipped_outside_scope": 0,
            "skipped_dzialka_exists": 0,
            "skipped_pe_exists": 0,
            "skipped_md_split_exists": 0,
            "skipped_spl-i-rz_exists": 0,
            "skipped_no_dzialka_found": 0,
            "skipped_no_kabel_found": 0,
            "skipped_no_dzialka_found_ids": [],
            "skipped_no_kabel_found_ids": [],
        }

        pa_layer.startEditing()

        dzialki_layer = self.dzialki_layer_combobox.currentData() if self.groupBox_dzialka.isChecked() else None
        dzialki_attr = self.dzialki_attribute_combobox.currentText() if self.groupBox_dzialka.isChecked() else None
        kable_layer_list = QgsProject.instance().mapLayersByName("kable")
        kable_layer = kable_layer_list[0] if kable_layer_list and self.groupBox_zasilanie.isChecked() else None

        dzialki_index = QgsSpatialIndex(dzialki_layer.getFeatures()) if dzialki_layer else None
        
        kabel_last_vertex_index = {}
        if kable_layer:
            self.output_widget.log_info("Filtrowanie kabli w zakresie i tworzenie indeksu wierzchołków...")
            # Użyj QgsFeatureRequest z prostokątem otaczającym, aby przyspieszyć filtrowanie
            request_kable = QgsFeatureRequest().setFilterRect(scope_geom.boundingBox())
            # Iteruj po kablach, które znajdują się w prostokącie otaczającym zakresu
            for kabel_feature in kable_layer.getFeatures(request_kable):
                kabel_geom = kabel_feature.geometry()

                if not kabel_geom or not QgsWkbTypes.geometryType(kabel_geom.wkbType()) == QgsWkbTypes.LineGeometry:
                    continue

                last_vertex = None
                # Correctly handle both single and multi-part line geometries
                if kabel_geom.isMultipart():
                    multi_line = kabel_geom.asMultiPolyline()
                    if multi_line and multi_line[-1]:
                        last_vertex = multi_line[-1][-1]
                else:
                    line = kabel_geom.asPolyline()
                    if line:
                        last_vertex = line[-1]
                
                if last_vertex:
                    # Check if the cable's last vertex is within the selected scope polygon
                    if QgsGeometry.fromPointXY(last_vertex).within(scope_geom):
                        rounded_coords = (round(last_vertex.x(), 8), round(last_vertex.y(), 8))
                        kabel_last_vertex_index[rounded_coords] = kabel_feature
            self.output_widget.log_info(f"Znaleziono {len(kabel_last_vertex_index)} pasujących kabli w zakresie.")

        transform = None
        if dzialki_layer and dzialki_layer.crs() != pa_layer.crs():
            transform = QgsCoordinateTransform(pa_layer.crs(), dzialki_layer.crs(), QgsProject.instance())

        request = QgsFeatureRequest().setFilterRect(scope_geom.boundingBox())
        for pa_feature in pa_layer.getFeatures(request):
            stats["processed"] += 1
            pa_geom = pa_feature.geometry()

            if not pa_geom:
                stats["skipped_no_geom"] += 1
                continue
            
            if not pa_geom.intersects(scope_geom):
                stats["skipped_outside_scope"] += 1
                continue

            if self.groupBox_dzialka.isChecked():
                current_dzialka = pa_feature["X_dzialka"]
                if overwrite or not current_dzialka:
                    pa_geom_transformed = QgsGeometry(pa_geom)
                    if transform:
                        pa_geom_transformed.transform(transform)
                    
                    intersecting_dzialki_ids = dzialki_index.intersects(pa_geom_transformed.boundingBox())
                    found_dzialka = False
                    for dzialka_id in intersecting_dzialki_ids:
                        dzialka_feature = dzialki_layer.getFeature(dzialka_id)
                        if dzialka_feature.geometry().contains(pa_geom_transformed):
                            new_dzialka_val = dzialka_feature[dzialki_attr]
                            pa_layer.changeAttributeValue(pa_feature.id(), pa_layer.fields().indexOf("X_dzialka"), new_dzialka_val)
                            stats["dzialka_assigned"] += 1
                            found_dzialka = True
                            break
                    if not found_dzialka:
                        stats["skipped_no_dzialka_found"] += 1
                        stats["skipped_no_dzialka_found_ids"].append(self._get_feature_identifier(pa_feature))
                else:
                    stats["skipped_dzialka_exists"] += 1

            if self.groupBox_zasilanie.isChecked():
                pa_point = pa_geom.asPoint()
                rounded_pa_coords = (round(pa_point.x(), 8), round(pa_point.y(), 8))
                
                kabel_feature = kabel_last_vertex_index.get(rounded_pa_coords)
                
                if kabel_feature:
                    if self.cb_przypisz_pe.isChecked():
                        current_pe = pa_feature["X_PE"]
                        if overwrite or not current_pe:
                            new_pe_val = kabel_feature["pe_poczatk"]
                            pa_layer.changeAttributeValue(pa_feature.id(), pa_layer.fields().indexOf("X_PE"), new_pe_val)
                            stats["pe_assigned"] += 1
                            stats["pe_values"][new_pe_val] += 1
                        else:
                            stats["skipped_pe_exists"] += 1

                    if self.cb_przypisz_md_split.isChecked():
                        current_md_split = pa_feature["X_MD_SPLIT"]
                        if overwrite or not current_md_split:
                            new_md_split_val = kabel_feature["X_MD_SPLIT"]
                            pa_layer.changeAttributeValue(pa_feature.id(), pa_layer.fields().indexOf("X_MD_SPLIT"), new_md_split_val)
                            stats["md_split_assigned"] += 1
                            stats["md_split_values"][new_md_split_val] += 1
                        else:
                            stats["skipped_md_split_exists"] += 1

                    if self.cb_przypisz_spl_i_rz.isChecked():
                        current_spl_i_rz = pa_feature["X_SPL-i-rz"]
                        if overwrite or not current_spl_i_rz:
                            new_spl_i_rz_val = kabel_feature["X_SPL-i-rz"]
                            pa_layer.changeAttributeValue(pa_feature.id(), pa_layer.fields().indexOf("X_SPL-i-rz"), new_spl_i_rz_val)
                            stats["spl-i-rz_assigned"] += 1
                            stats["spl-i-rz_values"][new_spl_i_rz_val] += 1
                        else:
                            stats["skipped_spl-i-rz_exists"] += 1
                else:
                    stats["skipped_no_kabel_found"] += 1
                    stats["skipped_no_kabel_found_ids"].append(self._get_feature_identifier(pa_feature))

        pa_layer.commitChanges()
        self._log_summary(stats)

    def _is_valid_for_run(self, pa_layer):
        if pa_layer.isEditable():
            self.output_widget.log_error(f"Warstwa '{pa_layer.name()}' jest w trybie edycji. Wyłącz tryb edycji, aby kontynuować.")
            return False

        pa_fields = pa_layer.fields().names()
        required_pa_id_fields = ["Id_budynku", "Miejscowos", "Ulica", "Numer porz"]
        for field in required_pa_id_fields:
            if field not in pa_fields:
                self.output_widget.log_error(f"Warstwa 'lista_pa' nie posiada wymaganego atrybutu identyfikacyjnego: '{field}'.")
                return False

        if self.groupBox_dzialka.isChecked():
            if not self.dzialki_layer_combobox.currentData():
                self.output_widget.log_error("Nie wybrano warstwy z działkami.")
                return False
            if not self.dzialki_attribute_combobox.currentText():
                self.output_widget.log_error("Nie wybrano atrybutu z numerem działki.")
                return False
            if "X_dzialka" not in pa_fields:
                self.output_widget.log_error("Warstwa 'lista_pa' nie posiada atrybutu 'X_dzialka'.")
                return False

        if self.groupBox_zasilanie.isChecked():
            if not self.cb_przypisz_pe.isChecked() and not self.cb_przypisz_md_split.isChecked() and not self.cb_przypisz_spl_i_rz.isChecked():
                self.output_widget.log_error("W sekcji 'Zasilanie' nie wybrano żadnej opcji do wykonania.")
                return False

            kable_layer_list = QgsProject.instance().mapLayersByName("kable")
            if not kable_layer_list:
                self.output_widget.log_error("Nie znaleziono warstwy 'kable'.")
                return False
            kable_fields = kable_layer_list[0].fields().names()

            if self.cb_przypisz_pe.isChecked():
                if "pe_poczatk" not in kable_fields:
                    self.output_widget.log_error("Warstwa 'kable' nie posiada atrybutu 'pe_poczatk'.")
                    return False
                if "X_PE" not in pa_fields:
                    self.output_widget.log_error("Warstwa 'lista_pa' nie posiada atrybutu 'X_PE'.")
                    return False
            
            if self.cb_przypisz_md_split.isChecked():
                if "X_MD_SPLIT" not in kable_fields:
                    self.output_widget.log_error("Warstwa 'kable' nie posiada atrybutu 'X_MD_SPLIT'.")
                    return False
                if "X_MD_SPLIT" not in pa_fields:
                    self.output_widget.log_error("Warstwa 'lista_pa' nie posiada atrybutu 'X_MD_SPLIT'.")
                    return False

            if self.cb_przypisz_spl_i_rz.isChecked():
                if "X_SPL-i-rz" not in kable_fields:
                    self.output_widget.log_error("Warstwa 'kable' nie posiada atrybutu 'X_SPL-i-rz'.")
                    return False
                if "X_SPL-i-rz" not in pa_fields:
                    self.output_widget.log_error("Warstwa 'lista_pa' nie posiada atrybutu 'X_SPL-i-rz'.")
                    return False
        
        if not self.groupBox_dzialka.isChecked() and not self.groupBox_zasilanie.isChecked():
            self.output_widget.log_error("Żadna sekcja nie została włączona. Zaznacz 'Przypisz działkę do PA' lub 'Przypisz zasilanie PA'.")
            return False

        return True

    def _log_summary(self, stats):
        self.output_widget.log_info("--- PODSUMOWANIE ---")
        self.output_widget.log_info(f"Łącznie przetworzono obiektów z warstwy 'lista_pa': {stats['processed']}")

        if self.groupBox_dzialka.isChecked():
            self.output_widget.log_info("--- Działki ---")
            self.output_widget.log_success(f"Przypisano numerów działek: {stats['dzialka_assigned']}")
            if not self.nadpisz_radio.isChecked():
                self.output_widget.log_warning(f"Pominięto (istniejąca wartość): {stats['skipped_dzialka_exists']}")
            if stats['skipped_no_dzialka_found_ids']:
                self.output_widget.log_warning(f"Nie znaleziono działki dla: {stats['skipped_no_dzialka_found']} obiektów")
                self.output_widget.log_info('\n'.join(stats['skipped_no_dzialka_found_ids']))

        if self.groupBox_zasilanie.isChecked():
            self.output_widget.log_info("--- Zasilanie ---")
            if self.cb_przypisz_pe.isChecked():
                self.output_widget.log_success(f"Przypisano wartości atrybutu X_PE: {stats['pe_assigned']}")
                if stats["pe_values"]:
                    self.output_widget.log_info("Przypisane wartości X_PE:")
                    for val, count in sorted(stats["pe_values"].items()):
                        self.output_widget.log_info(f"  - '{val}': {count} razy")
                if not self.nadpisz_radio.isChecked():
                    self.output_widget.log_warning(f"Pominięto X_PE (istniejąca wartość): {stats['skipped_pe_exists']}")
            
            if self.cb_przypisz_md_split.isChecked():
                self.output_widget.log_success(f"Przypisano wartości atrybutu X_MD_SPLIT: {stats['md_split_assigned']}")
                if stats["md_split_values"]:
                    self.output_widget.log_info("Przypisane wartości X_MD_SPLIT:")
                    for val, count in sorted(stats["md_split_values"].items()):
                        self.output_widget.log_info(f"  - '{val}': {count} razy")
                if not self.nadpisz_radio.isChecked():
                    self.output_widget.log_warning(f"Pominięto X_MD_SPLIT (istniejąca wartość): {stats['skipped_md_split_exists']}")

            if self.cb_przypisz_spl_i_rz.isChecked():
                self.output_widget.log_success(f"Przypisano wartości atrybutu X_SPL-i-rz: {stats['spl-i-rz_assigned']}")
                if stats["spl-i-rz_values"]:
                    self.output_widget.log_info("Przypisane wartości X_SPL-i-rz:")
                    for val, count in sorted(stats["spl-i-rz_values"].items()):
                        self.output_widget.log_info(f"  - '{val}': {count} razy")
                if not self.nadpisz_radio.isChecked():
                    self.output_widget.log_warning(f"Pominięto X_SPL-i-rz (istniejąca wartość): {stats['skipped_spl-i-rz_exists']}")

            if stats['skipped_no_kabel_found_ids']:
                self.output_widget.log_warning(f"Nie znaleziono kabla dla: {stats['skipped_no_kabel_found']} obiektów")
                self.output_widget.log_info('\n'.join(stats['skipped_no_kabel_found_ids']))

        if stats["skipped_no_geom"] > 0:
            self.output_widget.log_warning(f"Pominięto obiektów bez geometrii: {stats['skipped_no_geom']}")
        if stats["skipped_outside_scope"] > 0:
            self.output_widget.log_warning(f"Pominięto obiektów poza zakresem: {stats['skipped_outside_scope']}")
        
        self.output_widget.log_success("Zakończono pomyślnie.")

    def _browse_file_mr(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Wybierz plik Excel", "", "Pliki Excel (*.xlsx)")
        if file_path:
            self.file_path_line_edit_mr.setText(file_path)

    def run_aktualizuj_mr_action(self):
        self.output_widget.log_info("Uruchomiono 'Aktualizuj MR'...")

        pa_layer_list = QgsProject.instance().mapLayersByName("lista_pa")
        if not pa_layer_list:
            self.output_widget.log_error("Nie znaleziono warstwy 'lista_pa'.")
            return
        pa_layer = pa_layer_list[0]

        if not self._validate_aktualizuj_mr(pa_layer):
            self.output_widget.log_error("Walidacja nie powiodła się. Przerwana operacja.")
            return

        self.output_widget.log_success("Walidacja pomyślna. Rozpoczynanie operacji...")

        file_path = self.file_path_line_edit_mr.text()
        excel_data, stats = self._load_mr_data_from_excel(file_path)
        if not excel_data:
            return # Błąd został już zalogowany w _load_mr_data_from_excel

        excel_ids_to_match = set(excel_data.keys())
        scope_geom = self.zakres_combo_box.currentData()
        overwrite = self.nadpisz_radio_mr.isChecked()
        
        pa_layer.startEditing()
        mr_field_idx = pa_layer.fields().indexOf("MR")
        rodzaj_field_idx = pa_layer.fields().indexOf("Rodzaj pun")

        request = QgsFeatureRequest().setFilterRect(scope_geom.boundingBox())
        for feature in pa_layer.getFeatures(request):
            if not feature.geometry().intersects(scope_geom):
                continue

            stats["processed_pa"] += 1
            id_budynku = feature.attribute("Id_budynku")

            if id_budynku and id_budynku in excel_data:
                stats["matched"] += 1
                excel_ids_to_match.discard(id_budynku)
                item = excel_data[id_budynku]
                
                # 1. Aktualizacja MR
                new_mr_value = self._parse_and_format_mr(item["mr"], stats, item.get("row_idx"))
                
                if new_mr_value:
                    current_mr_value = feature.attribute("MR")
                    if overwrite or not current_mr_value:
                        if str(current_mr_value) != str(new_mr_value):
                            pa_layer.changeAttributeValue(feature.id(), mr_field_idx, new_mr_value)
                            stats["mr_value_changed"] += 1
                        else:
                            stats["mr_value_not_changed"] += 1
                        stats["mr_copied"] += 1
                    else:
                        stats["skipped"] += 1
                
                # 2. Porównanie rodzaju punktu
                rodzaj_qgis = (feature.attribute("Rodzaj pun") or "").lower()
                rodzaj_excel = (item["rodzaj"] or "").lower()

                is_qgis_podstawowy = "podstawowy" in rodzaj_qgis
                is_excel_podstawowy = "podstawowy" in rodzaj_excel

                if is_qgis_podstawowy != is_excel_podstawowy:
                    stats["rodzaj_discrepancy"] += 1
                    row_info = f" (wiersz: {item.get('row_idx')})" if item.get('row_idx') else ""
                    self.output_widget.log_warning(
                        f"UWAGA: Obiekt o Id_budynku: {id_budynku}{row_info} ma w projekcie rodzaj '{rodzaj_qgis}' a w pliku '{item['rodzaj']}'"
                    )

        pa_layer.commitChanges()
        
        stats["unmatched_excel"] = len(excel_ids_to_match)
        if excel_ids_to_match:
            self.output_widget.log_warning(f"Nie znaleziono dopasowania w warstwie 'lista_pa' dla {len(excel_ids_to_match)} obiektów z pliku Excel:")
            self.output_widget.log_info(", ".join(sorted(list(excel_ids_to_match))))

        self._log_mr_summary(stats)

    def _validate_aktualizuj_mr(self, pa_layer):
        if not openpyxl:
            self.output_widget.log_error("Biblioteka 'openpyxl' nie jest zainstalowana. Użyj polecenia 'pip install openpyxl' w konsoli OSGeo4W, aby ją zainstalować.")
            return False
        
        file_path = self.file_path_line_edit_mr.text()
        if not file_path or not os.path.exists(file_path):
            self.output_widget.log_error("Podana ścieżka do pliku jest nieprawidłowa lub plik nie istnieje.")
            return False

        if pa_layer.isEditable():
            self.output_widget.log_error(f"Warstwa '{pa_layer.name()}' jest w trybie edycji. Wyłącz tryb edycji, aby kontynuować.")
            return False

        pa_fields = pa_layer.fields().names()
        required_fields = ["Id_budynku", "MR", "Rodzaj pun"]
        for field in required_fields:
            if field not in pa_fields:
                self.output_widget.log_error(f"Warstwa '{pa_layer.name()}' nie posiada wymaganego atrybutu: '{field}'.")
                return False
        return True

    def _load_mr_data_from_excel(self, file_path):
        stats = defaultdict(int)
        data = {}
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            header = [cell.value for cell in sheet[1]]
            required_cols = ["Id_budynku", "MR", "Rodzaj"]
            
            col_indices = {}
            for col_name in required_cols:
                try:
                    col_indices[col_name] = header.index(col_name)
                except ValueError:
                    self.output_widget.log_error(f"W pliku Excel brakuje wymaganej kolumny: '{col_name}'.")
                    return None, None

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                stats["processed_excel"] += 1
                id_budynku = row[col_indices["Id_budynku"]]
                if id_budynku:
                    id_budynku = str(id_budynku).strip()
                    data[id_budynku] = {
                        "mr": row[col_indices["MR"]],
                        "rodzaj": row[col_indices["Rodzaj"]],
                        "row_idx": row_idx
                    }
                else:
                    self.output_widget.log_warning(f"Pominięto wiersz {row_idx} z pliku Excel z powodu braku wartości w kolumnie 'Id_budynku'.")

        except Exception as e:
            self.output_widget.log_error(f"Wystąpił błąd podczas odczytu pliku Excel: {e}")
            return None, None
        
        return data, stats

    def _parse_and_format_mr(self, raw_mr_value, stats, row_idx=None):
        if not raw_mr_value:
            return None
        
        val = str(raw_mr_value).strip()
        
        # Wzorzec Z1.KM1.MR1
        match1 = re.search(r'\.MR(\d+)$', val, re.IGNORECASE)
        if match1:
            stats["matched_pattern_Z1"] += 1
            num = int(match1.group(1))
            return f"MR{num:02d}"

        # Wzorzec M.HP.01.P
        match2 = re.search(r'^\w\.\w+\.(\d+)\.\w$', val, re.IGNORECASE)
        if match2:
            stats["matched_pattern_MHP"] += 1
            num = int(match2.group(1))
            return f"MR{num:02d}"

        # Wzorzec MRxx lub MRx
        match3 = re.search(r'^MR(\d{1,2})$', val, re.IGNORECASE)
        if match3:
            stats["matched_pattern_MRx"] += 1
            num = int(match3.group(1))
            return f"MR{num:02d}"

        row_info = f" (wiersz: {row_idx})" if row_idx else ""
        self.output_widget.log_warning(f"Nierozpoznany format wartości MR: '{val}'{row_info}. Wartość została pominięta.")
        return None

    def _log_mr_summary(self, stats):
        self.output_widget.log_info("--- PODSUMOWANIE: AKTUALIZACJA MR ---")
        self.output_widget.log_info(f"Ilość łącznie przetworzonych obiektów z warstwy 'lista_pa': {stats['processed_pa']}")
        self.output_widget.log_info(f"Ilość łącznie przetworzonych obiektów z pliku excel: {stats['processed_excel']}")
        self.output_widget.log_info(f"Ilość obiektów łącznie dla których znaleziono dopasowanie: {stats['matched']}")
        self.output_widget.log_warning(f"Ilość obiektów z pliku Excel, dla których nie znaleziono dopasowania w warstwie: {stats.get('unmatched_excel', 0)}")
        self.output_widget.log_info(f"Ilość obiektów dla których znaleziono dopasowanie ze wzorcem M.HP.xx.P: {stats['matched_pattern_MHP']}")
        self.output_widget.log_info(f"Ilość obiektów dla których znaleziono dopasowanie ze wzorcem Z1.KM1.xxxx: {stats['matched_pattern_Z1']}")
        self.output_widget.log_info(f"Ilość obiektów dla których znaleziono dopasowanie ze wzorcem MRx(x): {stats['matched_pattern_MRx']}")
        self.output_widget.log_success(f"Ilość obiektów dla których prawidłowo skopiowano wartość atrybutu MR: {stats['mr_copied']}")
        self.output_widget.log_success(f"Ilość obiektów dla których wartość atrybutu MR została zmieniona względem wcześniejszej: {stats['mr_value_changed']}")
        self.output_widget.log_info(f"Ilość obiektów dla których wartość atrybutu MR nie została zmieniona (była już prawidłowa): {stats['mr_value_not_changed']}")
        self.output_widget.log_warning(f"Ilość obiektów z rozbieżnością rodzaju punktu podstawowy/dodatkowy: {stats['rodzaj_discrepancy']}")
        if not self.nadpisz_radio_mr.isChecked():
            self.output_widget.log_warning(f"Ilość obiektów pominiętych (istniała już wartość w polu MR): {stats['skipped']}")
        self.output_widget.log_success("Zakończono pomyślnie.")


    def _get_identifiers(self):
        if self.tabWidget_input.currentIndex() == 0: # Plik
            return self._get_identifiers_from_file()
        else: # Ręcznie
            return self._get_identifiers_from_text()

    def _get_identifiers_from_file(self):
        file_path = self.file_path_line_edit.text()
        if not file_path:
            self.output_widget.log_error("Nie wybrano pliku.")
            return [], []

        raw_identifiers = []
        if file_path.endswith('.csv'):
            try:
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    header = next(reader, None)
                    # Sprawdzenie, czy plik ma jedną kolumnę, czy wartości po przecinku w jednym wierszu
                    if header and len(header) == 1:
                        raw_identifiers.append(header[0]) # Pierwszy wiersz, pierwsza kolumna
                        for row in reader:
                            if row:
                                raw_identifiers.append(row[0])
                    elif header:
                        raw_identifiers.extend(header)
                        for row in reader:
                            raw_identifiers.extend(row)

            except Exception as e:
                self.output_widget.log_error(f"Błąd odczytu pliku CSV: {e}")
                return [], []
        elif file_path.endswith('.xlsx'):
            if not openpyxl:
                self.output_widget.log_error("Biblioteka 'openpyxl' nie jest zainstalowana. Użyj pliku CSV lub zainstaluj bibliotekę (pip install openpyxl).")
                return [], []
            try:
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                for row in sheet.iter_rows(values_only=True):
                    for cell_value in row:
                        if cell_value is not None:
                            raw_identifiers.append(str(cell_value))
            except Exception as e:
                self.output_widget.log_error(f"Błąd odczytu pliku XLSX: {e}")
                return [], []
        
        cleaned_identifiers = [str(val).strip() for val in raw_identifiers if str(val).strip()]
        counts = defaultdict(int)
        for item in cleaned_identifiers:
            counts[item] += 1
        
        duplicates = {item: count for item, count in counts.items() if count > 1}
        return list(counts.keys()), duplicates

    def _get_identifiers_from_text(self):
        text = self.manual_input_text_edit.toPlainText()
        if not text:
            return [], []
        # Split by comma, semicolon or newline
        raw_identifiers = re.split(r'[,;\n]+', text)
        cleaned_identifiers = [val.strip() for val in raw_identifiers if val.strip()]
        
        counts = defaultdict(int)
        for item in cleaned_identifiers:
            counts[item] += 1
            
        duplicates = {item: count for item, count in counts.items() if count > 1}
        return list(counts.keys()), duplicates

    def _validate_wykluczanie_pa(self, pa_layer):
        if pa_layer.isEditable():
            self.output_widget.log_error(f"Warstwa '{pa_layer.name()}' jest w trybie edycji. Wyłącz tryb edycji, aby kontynuować.")
            return False

        pa_fields = pa_layer.fields().names()
        # Atrybuty potrzebne do logowania i ogólnej identyfikacji
        required_pa_id_fields = ["Id_budynku", "Miejscowos", "Ulica", "Numer porz"]
        for field in required_pa_id_fields:
            if field not in pa_fields:
                self.output_widget.log_error(f"Warstwa 'lista_pa' nie posiada wymaganego atrybutu identyfikacyjnego: '{field}'.")
                return False

        # Atrybuty specyficzne dla tej funkcjonalności
        required_attrs = ["X_dubel", "X_weryfik"]
        if self.radio_id_budynku.isChecked():
            required_attrs.append("Id_budynku")
        else:
            required_attrs.append("X_ID_TOK")

        for attr in required_attrs:
            if attr not in pa_fields:
                self.output_widget.log_error(f"Warstwa 'lista_pa' nie posiada wymaganego atrybutu: '{attr}'.")
                return False

        if self.tabWidget_input.currentIndex() == 0 and not self.file_path_line_edit.text():
            self.output_widget.log_error("Nie wybrano pliku z identyfikatorami.")
            return False
        
        if self.tabWidget_input.currentIndex() == 1 and not self.manual_input_text_edit.toPlainText():
            self.output_widget.log_error("Nie wprowadzono ręcznie żadnych identyfikatorów.")
            return False

        return True

    def _log_wykluczanie_summary(self, stats):
        self.output_widget.log_info("--- PODSUMOWANIE WYKLUCZANIA PA ---")
        self.output_widget.log_info(f"Łącznie przetworzono obiektów z warstwy 'lista_pa': {stats['processed']}")
        
        total_user_ids = len(stats['unmatched_ids']) + stats['matched_id_budynku'] + stats['matched_x_id_tok']
        self.output_widget.log_info(f"Łącznie poszukiwanych unikalnych identyfikatorów: {total_user_ids}")

        if stats['user_duplicates']:
            self.output_widget.log_warning(f"Znaleziono {len(stats['user_duplicates'])} zdublowanych wartości w danych wejściowych:")
            for val, count in stats['user_duplicates'].items():
                self.output_widget.log_warning(f"- '{val}' (wystąpiło {count} razy)")

        self.output_widget.log_info(f"Dopasowano obiektów na podstawie 'Id_budynku': {stats['matched_id_budynku']}")
        self.output_widget.log_info(f"Dopasowano obiektów na podstawie 'X_ID_TOK': {stats['matched_x_id_tok']}")
        self.output_widget.log_success(f"Ustawiono 'X_dubel' na 'TAK' dla: {stats['x_dubel_set']} obiektów")
        self.output_widget.log_success(f"Ustawiono 'X_weryfik' na 'TAK' dla: {stats['x_weryfik_set']} obiektów")
        
        if not self.nadpisz_radio_wykluczanie.isChecked():
            self.output_widget.log_warning(f"Pominięto (istniejąca wartość): {stats['skipped_existing']} obiektów")

        if stats['unmatched_ids']:
            self.output_widget.log_error(f"Nie znaleziono dopasowania dla {len(stats['unmatched_ids'])} identyfikatorów:")
            self.output_widget.log_info(", ".join(sorted(list(stats['unmatched_ids']))))

        self.output_widget.log_success("Zakończono pomyślnie.")