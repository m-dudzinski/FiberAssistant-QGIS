import os
import json
import csv
import traceback
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None

from qgis.PyQt import uic
from qgis.PyQt.QtWidgets import QWidget, QVBoxLayout, QFileDialog
from qgis.core import QgsProject, QgsFeatureRequest, QgsSpatialIndex

from ..core.logger import logger
from .base_widget import FormattedOutputWidget

FORM_CLASS, _ = uic.loadUiType(os.path.join(
    os.path.dirname(__file__), '../ui/karta_krosowan_widget.ui'))

class KartaKrosowanWidget(QWidget, FORM_CLASS):
    FUNCTIONALITY_NAME = "Karta krosowań"

    def __init__(self, iface, parent=None):
        super(KartaKrosowanWidget, self).__init__(parent)
        self.iface = iface
        self.logger = logger
        self.setupUi(self)
        self.loaded_data = None

        self._setup_output_widget()
        self._connect_signals()
        self._setup_initial_state()
        self._populate_scope_combobox()
        
        self.splitter.setSizes([400, 150])

    def _setup_output_widget(self):
        self.output_widget = FormattedOutputWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.output_widget)
        self.output_widget_placeholder.setLayout(layout)
        self.logger.set_user_message_widget(self.output_widget.output_console)

    def _connect_signals(self):
        self.browse_button.clicked.connect(self._browse_file)
        self.refresh_button.clicked.connect(self._populate_scope_combobox)
        self.radio_scope_selected.toggled.connect(self._toggle_scope_widgets)

    def _toggle_scope_widgets(self, enabled):
        self.zakres_combo_box.setEnabled(enabled)
        self.refresh_button.setEnabled(enabled)
        self.label_zakres.setEnabled(enabled)

    def _setup_initial_state(self):
        self.groupBox_strategy.setEnabled(False)

    def _populate_scope_combobox(self):
        self.zakres_combo_box.clear()
        zakres_layer_list = QgsProject.instance().mapLayersByName("zakres_zadania")
        if not zakres_layer_list:
            self.output_widget.log_error("Nie znaleziono warstwy 'zakres_zadania'.")
            return
        
        zakres_layer = zakres_layer_list[0]
        if "nazwa" not in zakres_layer.fields().names():
            self.output_widget.log_error("Warstwa 'zakres_zadania' nie posiada atrybutu 'nazwa'.")
            return

        features_to_add = []
        for feature in zakres_layer.getFeatures():
            if feature.attribute("nazwa"):
                features_to_add.append((feature["nazwa"], feature.geometry()))

        features_to_add.sort(key=lambda f: f[0])

        for nazwa, geom in features_to_add:
            self.zakres_combo_box.addItem(nazwa, geom)

        self.output_widget.log_info(f"Znaleziono {self.zakres_combo_box.count()} zakresów.")

    def _browse_file(self):
        self.logger.log_dev(self.FUNCTIONALITY_NAME, 0, "USER", "Kliknięto przycisk 'Przeglądaj...'")
        file_path, _ = QFileDialog.getOpenFileName(self, "Wybierz plik karty krosowań", "", "Pliki Excel (*.xlsx);;Pliki CSV (*.csv)")
        if file_path:
            self.file_path_line_edit.setText(file_path)
            self.output_widget.clear_log()
            self.logger.log_user(f"Wybrano plik: {file_path}")
            self._load_and_validate_file(file_path)

    def _load_and_validate_file(self, file_path):
        self.loaded_data = None
        self.groupBox_strategy.setEnabled(False)
        data = None
        error = None

        if file_path.endswith('.xlsx'):
            if not openpyxl:
                self.output_widget.log_error("Biblioteka 'openpyxl' nie jest zainstalowana. Użyj polecenia 'pip install openpyxl' w konsoli OSGeo4W.")
                return

            workbook = None
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                data = self._extract_data_from_workbook(workbook)
            except Exception as e:
                error = f"Błąd podczas odczytu pliku XLSX: {e}"
            finally:
                if workbook:
                    workbook.close()
        
        elif file_path.endswith('.csv'):
            data, error = self._read_csv(file_path)
        else:
            self.output_widget.log_error("Nieobsługiwany format pliku. Wybierz plik .xlsx lub .csv.")
            return

        if error:
            self.output_widget.log_error(error)
            return

        if not data:
            self.output_widget.log_warning("Plik nie zawiera żadnych danych.")
            return

        is_valid, validation_error, validation_warning = self._validate_data_structure(data)
        if not is_valid:
            self.output_widget.log_error(validation_error)
            return

        self.output_widget.log_success("Wczytano kartę krosowań zgodną ze wzorcem.")
        if validation_warning:
             self.output_widget.log_warning(validation_warning)

        self.loaded_data = data
        self.groupBox_strategy.setEnabled(True)
        self.output_widget.log_info(f"Pomyślnie załadowano i zweryfikowano {len(self.loaded_data)} wierszy danych.")

    def _extract_data_from_workbook(self, workbook):
        data = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            header = [cell.value for cell in sheet[1]]
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if any(row):
                    row_data = dict(zip(header, row))
                    row_data['__sheet__'] = sheet_name
                    row_data['__row_idx__'] = row_idx
                    data.append(row_data)
        return data

    def _read_csv(self, file_path):
        data = []
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                try:
                    dialect = csv.Sniffer().sniff(f.read(1024))
                    f.seek(0)
                except csv.Error:
                    dialect = 'excel' 
                    f.seek(0)

                reader = csv.DictReader(f, dialect=dialect)
                for row_idx, row in enumerate(reader, start=2):
                    if any(row.values()):
                        row['__sheet__'] = os.path.basename(file_path)
                        row['__row_idx__'] = row_idx
                        data.append(row)
            return data, None
        except Exception as e:
            return None, f"Błąd podczas odczytu pliku CSV: {e}"

    def _validate_data_structure(self, data):
        try:
            wzorzec_path = os.path.join(os.path.dirname(__file__), "..", "templates", "karta_krosowan_wzorzec_zrzutu_danych_PA.json")
            with open(wzorzec_path, 'r', encoding='utf-8') as f:
                wzorzec = json.load(f)
            required_columns = wzorzec.get("required_columns", [])
        except Exception as e:
            return False, f"Nie można wczytać pliku wzorca '{os.path.basename(wzorzec_path)}': {e}", None

        if not data:
            return True, None, "Wczytano kartę krosowań zgodną ze wzorcem. UWAGA! Wykryto puste wiersze! Możliwe, że dane w karcie są niekompletne."

        header = data[0].keys()
        missing_cols = [col for col in required_columns if col not in header]
        if missing_cols:
            return False, f"Niewłaściwa struktura pliku. Brakuje kolumn: {', '.join(missing_cols)}", None

        empty_key_rows = []
        key_cols = ["X_port_olt", "X_SPL-i-rz", "X_MD_SPLIT"]
        for row in data:
            other_cols_have_data = any(row.get(col) for col in required_columns if col not in key_cols)
            key_cols_are_empty = not all(row.get(col) for col in key_cols)

            if other_cols_have_data and key_cols_are_empty:
                empty_key_rows.append(row.get('__row_idx__', 'N/A'))
        
        if empty_key_rows:
            warning_msg = (f"Wczytano kartę krosowań zgodną ze wzorcem. UWAGA! Wykryto puste komórki w kluczowych kolumnach "
                           f"('X_port_olt', 'X_SPL-i-rz', 'X_MD_SPLIT') w wierszach: {', '.join(map(str, empty_key_rows[:5]))}"
                           f"{'...' if len(empty_key_rows) > 5 else ''}. Dane mogą być niekompletne.")
            return True, None, warning_msg

        return True, None, None

    def run_main_action(self):
        current_tab_index = self.tabWidget.currentIndex()
        if current_tab_index == 0:
            self._run_action_przypisz_port()
        elif current_tab_index == 1:
            self._run_action_generuj_zrzut()
        else:
            self.output_widget.log_error("Nieznana zakładka.")

    def _run_action_przypisz_port(self):
        self.output_widget.clear_log()
        self.logger.log_user("Uruchomiono 'Przypisz port OLT do PA'...")
        
        if not self.loaded_data:
            self.output_widget.log_error("Nie wczytano żadnych danych z pliku. Przerwana operacja.")
            return

        if self.radio_scope_all.isChecked():
            scope_geom = None
            self.output_widget.log_info("Przetwarzanie dla całego projektu (bez ograniczenia zakresem).")
        else:
            scope_geom = self.zakres_combo_box.currentData()
            if not scope_geom or scope_geom.isEmpty():
                self.output_widget.log_error("Nie wybrano prawidłowego zakresu zadania.")
                return
            self.output_widget.log_info(f"Przetwarzanie dla zakresu: {self.zakres_combo_box.currentText()}")

        layers = self._get_layers([
            "punkty_elastycznosci", "zakres_splitera", "kable", "lista_pa"
        ])
        if not layers:
            return

        overwrite = self.radio_overwrite.isChecked()
        stats = defaultdict(lambda: defaultdict(int))

        pe_lookup = self._prepare_pe_lookup(self.loaded_data)

        self._update_punkty_elastycznosci(layers["punkty_elastycznosci"], pe_lookup, scope_geom, overwrite, stats)
        self._update_zakres_splitera(layers["zakres_splitera"], layers["punkty_elastycznosci"], scope_geom, overwrite, stats)
        self._update_kable(layers["kable"], layers["zakres_splitera"], scope_geom, overwrite, stats)
        self._update_lista_pa(layers["lista_pa"], layers["zakres_splitera"], scope_geom, overwrite, stats)
        
        self._log_summary(stats)

    def _run_action_generuj_zrzut(self):
        self.output_widget.clear_log()
        self.logger.log_user("Uruchomiono 'Generuj zrzut'...")

        if not pd:
            self.output_widget.log_error("Brak biblioteki 'pandas'. Użyj polecenia 'pip install pandas' w konsoli OSGeo4W.")
            return

        if self.radio_scope_all_zrzut.isChecked():
            scope_geom = None
            self.output_widget.log_info("Przetwarzanie dla całego projektu (bez ograniczenia zakresem).")
        else:
            scope_geom = self.zakres_combo_box.currentData()
            if not scope_geom or scope_geom.isEmpty():
                self.output_widget.log_error("Nie wybrano prawidłowego zakresu zadania.")
                return
            self.output_widget.log_info(f"Przetwarzanie dla zakresu: {self.zakres_combo_box.currentText()}")

        if self.radio_format_xlsx.isChecked():
            self._perform_xlsx_export(scope_geom)
        else:
            self._perform_csv_export(scope_geom)

    def _process_layers(self, scope_geom, stats):
        layers_to_process = {
            'PA': 'lista_pa',
            'PE': 'punkty_elastycznosci',
            'KABLE': 'kable'
        }
        for sheet_name, layer_name in layers_to_process.items():
            self.output_widget.log_info(f"Przetwarzanie warstwy '{layer_name}'...")
            
            layer_list = QgsProject.instance().mapLayersByName(layer_name)
            if not layer_list:
                self.output_widget.log_warning(f"Nie znaleziono warstwy '{layer_name}'.")
                stats[layer_name]['skipped_reason'] = "Nie znaleziono warstwy"
                yield layer_name, sheet_name, None
                continue
            
            layer = layer_list[0]
            
            request = QgsFeatureRequest()
            if scope_geom:
                request.setFilterRect(scope_geom.boundingBox())

            features_data = []
            field_names = layer.fields().names()
            
            for feature in layer.getFeatures(request):
                if scope_geom and not feature.geometry().intersects(scope_geom):
                    stats[layer_name]['skipped_not_in_scope'] += 1
                    continue
                
                stats[layer_name]['processed'] += 1
                
                if layer_name == 'lista_pa':
                    rodzaj_pun = feature.attribute("Rodzaj pun") or "BRAK"
                    stats[layer_name][f"typ_{rodzaj_pun}"] += 1
                elif layer_name == 'kable':
                    rodzaj = feature.attribute("rodzaj") or "BRAK"
                    stats[layer_name][f"typ_{rodzaj}"] += 1

                attrs = {field.name(): feature[field.name()] for field in layer.fields()}
                attrs['wkt_geom'] = feature.geometry().asWkt()
                features_data.append(attrs)

            if not features_data:
                self.output_widget.log_warning(f"Nie znaleziono obiektów w warstwie '{layer_name}' dla wybranego zakresu.")
                df = pd.DataFrame(columns=['wkt_geom'] + field_names)
            else:
                df = pd.DataFrame(features_data)
                df = df[['wkt_geom'] + field_names]
            
            yield layer_name, sheet_name, df

    def _perform_xlsx_export(self, scope_geom):
        if not xlsxwriter:
            self.output_widget.log_error(
                "Eksport do .xlsx wymaga biblioteki 'xlsxwriter', która nie jest zainstalowana.\n"
                "Zainstaluj ją (użyj polecenia 'pip install xlsxwriter' w konsoli OSGeo4W) "
                "lub użyj opcji eksportu do .csv."
            )
            return

        engine = 'xlsxwriter'
        file_path, _ = QFileDialog.getSaveFileName(self, "Zapisz zrzut danych do pliku Excel", "", "Plik Excel (*.xlsx)")
        if not file_path:
            self.output_widget.log_info("Operacja anulowana przez użytkownika.")
            return

        stats = defaultdict(lambda: defaultdict(int))
        try:
            with pd.ExcelWriter(file_path, engine=engine) as writer:
                for layer_name, sheet_name, df in self._process_layers(scope_geom, stats):
                    if df is None:
                        pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)
                        continue
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    self.output_widget.log_info(f"Zapisano {len(df)} obiektów do arkusza '{sheet_name}'.")
            
            self._log_dump_summary(stats)

        except Exception as e:
            self.output_widget.log_error(f"Wystąpił nieoczekiwany błąd podczas eksportu do XLSX: {e}")
            self.logger.log_dev(self.FUNCTIONALITY_NAME, 3, "ERROR", f"Błąd podczas generowania zrzutu XLSX: {e}\n{traceback.format_exc()}")

    def _perform_csv_export(self, scope_geom):
        file_path, _ = QFileDialog.getSaveFileName(self, "Wybierz lokalizację i bazową nazwę dla plików CSV", "", "Plik CSV (*.csv)")
        if not file_path:
            self.output_widget.log_info("Operacja anulowana przez użytkownika.")
            return

        base_path, _ = os.path.splitext(file_path)
        stats = defaultdict(lambda: defaultdict(int))
        
        try:
            for layer_name, sheet_name, df in self._process_layers(scope_geom, stats):
                if df is None or df.empty:
                    continue
                
                csv_path = f"{base_path}_{sheet_name}.csv"
                df.to_csv(csv_path, index=False, encoding='utf-8-sig', sep=';')
                self.output_widget.log_info(f"Zapisano {len(df)} obiektów do pliku '{os.path.basename(csv_path)}'.")

            self._log_dump_summary(stats)

        except Exception as e:
            self.output_widget.log_error(f"Wystąpił nieoczekiwany błąd podczas eksportu do CSV: {e}")
            self.logger.log_dev(self.FUNCTIONALITY_NAME, 3, "ERROR", f"Błąd podczas generowania zrzutu CSV: {e}\n{traceback.format_exc()}")

    def _log_dump_summary(self, stats):
        self.output_widget.log_info("\n--- PODSUMOWANIE GENEROWANIA ZRZUTU ---")
        
        pa_stats = stats.get('lista_pa', defaultdict(int))
        pa_total = pa_stats.get('processed', 0)
        self.output_widget.log_info(f"\nWarstwa: 'lista_pa'")
        self.output_widget.log_info(f"  - Ilość łącznie przetworzonych i dopasowanych obiektów: {pa_total}")
        if pa_total > 0:
            self.output_widget.log_info("  - Podział na 'Rodzaj pun':")
            for key, count in pa_stats.items():
                if key.startswith('typ_'):
                    self.output_widget.log_info(f"    - {key.replace('typ_', '')}: {count}")

        pe_stats = stats.get('punkty_elastycznosci', defaultdict(int))
        pe_total = pe_stats.get('processed', 0)
        self.output_widget.log_info(f"\nWarstwa: 'punkty_elastycznosci'")
        self.output_widget.log_info(f"  - Ilość łącznie przetworzonych i dopasowanych obiektów: {pe_total}")

        kable_stats = stats.get('kable', defaultdict(int))
        kable_total = kable_stats.get('processed', 0)
        self.output_widget.log_info(f"\nWarstwa: 'kable'")
        self.output_widget.log_info(f"  - Ilość łącznie przetworzonych i dopasowanych obiektów: {kable_total}")
        if kable_total > 0:
            self.output_widget.log_info("  - Podział na 'rodzaj':")
            for key, count in kable_stats.items():
                if key.startswith('typ_'):
                    self.output_widget.log_info(f"    - {key.replace('typ_', '')}: {count}")
        
        self.output_widget.log_info("\n--- Obiekty/Warstwy pominięte ---")
        any_skipped = False
        for layer_name, layer_stats in stats.items():
            reason = layer_stats.get('skipped_reason')
            if reason:
                self.output_widget.log_warning(f"  - Warstwa '{layer_name}': {reason}")
                any_skipped = True
            
            not_in_scope = layer_stats.get('skipped_not_in_scope')
            if not_in_scope:
                self.output_widget.log_warning(f"  - Warstwa '{layer_name}': {not_in_scope} obiektów poza zakresem.")
                any_skipped = True

        if not any_skipped:
            self.output_widget.log_info("  - Brak.")

        self.output_widget.log_success("\nZakończono pomyślnie generowanie zrzutu.")

    def _get_layers(self, layer_names):
        layers = {}
        project = QgsProject.instance()
        for name in layer_names:
            layer_list = project.mapLayersByName(name)
            if not layer_list:
                self.output_widget.log_error(f"Nie znaleziono wymaganej warstwy: '{name}'.")
                return None
            layer = layer_list[0]
            if layer.isEditable():
                self.output_widget.log_error(f"Warstwa '{name}' jest w trybie edycji. Wyłącz tryb edycji, aby kontynuować.")
                return None
            layers[name] = layer
        return layers

    def _prepare_pe_lookup(self, data):
        lookup = {}
        for row in data:
            pe_values = row.get("X_PE")
            if not pe_values:
                continue
            pe_list = [pe.strip() for pe in str(pe_values).split(',')]
            for pe_name in pe_list:
                if pe_name:
                    lookup[pe_name] = {
                        "X_SPL-i-rz": row.get("X_SPL-i-rz"),
                        "X_MD_SPLIT": row.get("X_MD_SPLIT"),
                        "X_port_olt": row.get("X_port_olt")
                    }
        return lookup

    def _update_punkty_elastycznosci(self, layer, pe_lookup, scope_geom, overwrite, stats):
        self.output_widget.log_info("Krok 1: Aktualizacja warstwy 'punkty_elastycznosci'...")
        valid_types = {"mufa", "szafka", "ODF", "skrzynka", "słupek"}
        attrs_to_update = ["X_SPL-i-rz", "X_MD_SPLIT", "X_port_olt"]
        
        layer.startEditing()
        
        request = QgsFeatureRequest()
        if scope_geom:
            request.setFilterRect(scope_geom.boundingBox())

        for feature in layer.getFeatures(request):
            if scope_geom and not feature.geometry().intersects(scope_geom):
                continue
            
            stats["punkty_elastycznosci"]["processed"] += 1
            
            if feature["typ"] not in valid_types:
                stats["punkty_elastycznosci"]["skipped_type"] += 1
                continue

            feature_name = feature["nazwa"]
            if not feature_name:
                stats["punkty_elastycznosci"]["skipped_no_name"] += 1
                continue

            matched_data = None
            for pe_name, data in pe_lookup.items():
                if feature_name in pe_name:
                    matched_data = data
                    break
            
            if not matched_data:
                stats["punkty_elastycznosci"]["skipped_no_match"] += 1
                continue

            for attr in attrs_to_update:
                current_val = feature[attr]
                new_val = matched_data.get(attr)
                
                if new_val is None:
                    continue

                if overwrite or not current_val:
                    if str(current_val) != str(new_val):
                        layer.changeAttributeValue(feature.id(), layer.fields().indexOf(attr), new_val)
                        stats["punkty_elastycznosci"][f"changed_{attr}"] += 1
                    else:
                        stats["punkty_elastycznosci"][f"not_changed_{attr}"] += 1
                    stats["punkty_elastycznosci"][f"assigned_{attr}"] += 1
                    stats[f"unique_{attr}"].setdefault(new_val, 0)
                    stats[f"unique_{attr}"][new_val] += 1
                else:
                    stats["punkty_elastycznosci"][f"skipped_existing_{attr}"] += 1
        
        layer.commitChanges()
        self.output_widget.log_info("Zakończono Krok 1.")

    def _update_zakres_splitera(self, zs_layer, pe_layer, scope_geom, overwrite, stats):
        self.output_widget.log_info("Krok 2: Aktualizacja warstwy 'zakres_splitera'...")
        attrs_to_update = ["X_port_olt", "X_SPL-i-rz", "X_MD_SPLIT"]
        
        pe_index = QgsSpatialIndex(pe_layer.getFeatures())
        
        zs_layer.startEditing()
        
        request = QgsFeatureRequest()
        if scope_geom:
            request.setFilterRect(scope_geom.boundingBox())

        for zs_feature in zs_layer.getFeatures(request):
            if scope_geom and not zs_feature.geometry().intersects(scope_geom):
                continue
            
            stats["zakres_splitera"]["processed"] += 1
            
            intersecting_pe_ids = pe_index.intersects(zs_feature.geometry().boundingBox())
            first_match_pe = None
            for pe_id in intersecting_pe_ids:
                pe_feature = pe_layer.getFeature(pe_id)
                if pe_feature.geometry().within(zs_feature.geometry()):
                    first_match_pe = pe_feature
                    break
            
            if not first_match_pe:
                stats["zakres_splitera"]["skipped_no_pe"] += 1
                continue

            for attr in attrs_to_update:
                current_val = zs_feature[attr]
                new_val = first_match_pe[attr]
                
                if new_val is None:
                    continue

                if overwrite or not current_val:
                    if str(current_val) != str(new_val):
                        zs_layer.changeAttributeValue(zs_feature.id(), zs_layer.fields().indexOf(attr), new_val)
                        stats["zakres_splitera"][f"changed_{attr}"] += 1
                    else:
                        stats["zakres_splitera"][f"not_changed_{attr}"] += 1
                    stats["zakres_splitera"][f"assigned_{attr}"] += 1
                else:
                    stats["zakres_splitera"][f"skipped_existing_{attr}"] += 1
            
            attr_split_i_rz = "split-i-rz"
            current_val_split_i_rz = zs_feature[attr_split_i_rz]
            new_val_split_i_rz = first_match_pe["X_SPL-i-rz"]
            if new_val_split_i_rz is not None:
                if overwrite or not current_val_split_i_rz:
                    if str(current_val_split_i_rz) != str(new_val_split_i_rz):
                        zs_layer.changeAttributeValue(zs_feature.id(), zs_layer.fields().indexOf(attr_split_i_rz), new_val_split_i_rz)
                        stats["zakres_splitera"][f"changed_{attr_split_i_rz}"] += 1
                    else:
                        stats["zakres_splitera"][f"not_changed_{attr_split_i_rz}"] += 1
                    stats["zakres_splitera"][f"assigned_{attr_split_i_rz}"] += 1
                else:
                    stats["zakres_splitera"][f"skipped_existing_{attr_split_i_rz}"] += 1

        zs_layer.commitChanges()
        self.output_widget.log_info("Zakończono Krok 2.")

    def _update_layer_from_zakres_splitera(self, layer, zs_layer, scope_geom, overwrite, stats, layer_name, feature_filter=None):
        self.output_widget.log_info(f"Krok {stats['step']}: Aktualizacja warstwy '{layer_name}'...")
        stats['step'] += 1
        attrs_to_update = ["X_port_olt", "X_SPL-i-rz", "X_MD_SPLIT"]
        
        zs_index = QgsSpatialIndex(zs_layer.getFeatures())
        
        layer.startEditing()
        
        request = QgsFeatureRequest()
        if scope_geom:
            request.setFilterRect(scope_geom.boundingBox())

        for feature in layer.getFeatures(request):
            if scope_geom and not feature.geometry().intersects(scope_geom):
                continue
            
            stats[layer_name]["processed"] += 1
            
            if feature_filter and not feature_filter(feature):
                stats[layer_name]["skipped_filter"] += 1
                continue

            intersecting_zs_ids = zs_index.intersects(feature.geometry().centroid().boundingBox())
            containing_zs = None
            for zs_id in intersecting_zs_ids:
                zs_feature = zs_layer.getFeature(zs_id)
                if zs_feature.geometry().contains(feature.geometry().centroid()):
                    containing_zs = zs_feature
                    break
            
            if not containing_zs:
                stats[layer_name]["skipped_no_zs"] += 1
                continue

            for attr in attrs_to_update:
                current_val = feature[attr]
                new_val = containing_zs[attr]
                
                if new_val is None:
                    continue

                if overwrite or not current_val:
                    if str(current_val) != str(new_val):
                        layer.changeAttributeValue(feature.id(), layer.fields().indexOf(attr), new_val)
                        stats[layer_name][f"changed_{attr}"] += 1
                    else:
                        stats[layer_name][f"not_changed_{attr}"] += 1
                    stats[layer_name][f"assigned_{attr}"] += 1
                else:
                    stats[layer_name][f"skipped_existing_{attr}"] += 1
        
        layer.commitChanges()
        self.output_widget.log_info(f"Zakończono Krok {stats['step']-1}.")

    def _update_kable(self, kable_layer, zs_layer, scope_geom, overwrite, stats):
        def kable_filter(feature):
            return (feature["segment"] == "abonencki" and 
                    feature["rodzaj"] in ["abonencki napowietrzny", "abonencki doziemny", "abonencki planowany"])
        
        stats['step'] = 3
        self._update_layer_from_zakres_splitera(kable_layer, zs_layer, scope_geom, overwrite, stats, "kable", kable_filter)

    def _update_lista_pa(self, pa_layer, zs_layer, scope_geom, overwrite, stats):
        stats['step'] = 4
        self._update_layer_from_zakres_splitera(pa_layer, zs_layer, scope_geom, overwrite, stats, "lista_pa")

    def _log_summary(self, stats):
        self.output_widget.log_info("--- PODSUMOWANIE ---")
        overwrite = self.radio_overwrite.isChecked()

        for layer_name in ["punkty_elastycznosci", "zakres_splitera", "kable", "lista_pa"]:
            if not stats[layer_name]: continue
            
            self.output_widget.log_info(f"\nWarstwa: '{layer_name}'")
            self.output_widget.log_info(f"  - Ilość łącznie przetworzonych obiektów: {stats[layer_name]['processed']}")
            
            attrs_to_log = ["X_port_olt", "X_SPL-i-rz", "X_MD_SPLIT"]
            if layer_name == "zakres_splitera":
                attrs_to_log.append("split-i-rz")

            for attr in attrs_to_log:
                assigned = stats[layer_name].get(f"assigned_{attr}", 0)
                if assigned > 0:
                    self.output_widget.log_success(f"  - Przypisano '{attr}': {assigned} obiektów")
                    if overwrite:
                        changed = stats[layer_name].get(f"changed_{attr}", 0)
                        not_changed = stats[layer_name].get(f"not_changed_{attr}", 0)
                        self.output_widget.log_info(f"    - Zmieniono wartość dla: {changed} obiektów")
                        self.output_widget.log_info(f"    - Wartość bez zmian dla: {not_changed} obiektów")
                
                skipped_existing = stats[layer_name].get(f"skipped_existing_{attr}", 0)
                if skipped_existing > 0:
                    self.output_widget.log_warning(f"  - Pominięto (istniejąca wartość) dla '{attr}': {skipped_existing} obiektów")

            if layer_name == "punkty_elastycznosci":
                if stats[layer_name]['skipped_type'] > 0: self.output_widget.log_warning(f"  - Pominięto (nieprawidłowy typ): {stats[layer_name]['skipped_type']}")
                if stats[layer_name]['skipped_no_name'] > 0: self.output_widget.log_warning(f"  - Pominięto (brak nazwy): {stats[layer_name]['skipped_no_name']}")
                if stats[layer_name]['skipped_no_match'] > 0: self.output_widget.log_warning(f"  - Pominięto (brak dopasowania w pliku): {stats[layer_name]['skipped_no_match']}")
            if layer_name == "zakres_splitera":
                if stats[layer_name]['skipped_no_pe'] > 0: self.output_widget.log_warning(f"  - Pominięto (brak PE wewnątrz): {stats[layer_name]['skipped_no_pe']}")
            if layer_name in ["kable", "lista_pa"]:
                if stats[layer_name]['skipped_filter'] > 0: self.output_widget.log_warning(f"  - Pominięto (niespełnione kryteria filtra): {stats[layer_name]['skipped_filter']}")
                if stats[layer_name]['skipped_no_zs'] > 0: self.output_widget.log_warning(f"  - Pominięto (brak zawierającego zakresu splitera): {stats[layer_name]['skipped_no_zs']}")

        self.output_widget.log_info("\n--- Unikalne wartości ---")
        for attr in ["X_port_olt", "X_SPL-i-rz", "X_MD_SPLIT"]:
            unique_values = stats.get(f"unique_{attr}")
            if unique_values:
                self.output_widget.log_info(f"Atrybut '{attr}' (ilość przypisań):")
                for value, count in sorted(unique_values.items()):
                    self.output_widget.log_info(f"  - {value}: {count}")
        
        self.output_widget.log_success("\nZakończono pomyślnie.")
        
    def refresh_data(self):
        self.output_widget.log_info("Odświeżanie listy zakresów...")
        self._populate_scope_combobox()
        self.output_widget.log_info("Lista zakresów została zaktualizowana.")
